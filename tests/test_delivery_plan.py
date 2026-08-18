import os
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

os.environ["DATABASE_PATH"] = "/tmp/selsa-planlama-feedback-test.sqlite"
os.environ["UPLOAD_DIR"] = "/tmp/selsa-planlama-feedback-uploads"
os.environ["ADMIN_PASSWORD"] = "test-password"
os.environ["APP_SESSION_SECRET"] = "test-session-secret-that-is-long-enough"

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.delivery_plan import build_delivery_plan
from app.main import app


def delivery_columns(count: int = 6):
    base = [
        {"label": "KW31\nGerçekleşen", "kind": "actual"},
        {"label": "KW32\nGerçekleşen", "kind": "actual"},
        {"label": "KW33\nGerçekleşen", "kind": "actual"},
        {"label": "KW33'e\nDevreden", "kind": "carryover"},
        {"label": "KW33\nPlan", "kind": "plan"},
        {"label": "KW34\nPlan", "kind": "plan"},
    ]
    return base + [{"label": f"KW{35 + index}\nPlan", "kind": "plan"} for index in range(max(0, count - len(base)))]


def test_delivery_plan_uses_single_category_and_template_structure():
    template = Path(__file__).resolve().parents[1] / "teslimat_plani.xlsx"
    content = build_delivery_plan(template, {
        "startDate": "2026-07-20",
        "endDate": "2026-08-30",
        "category": "Üretim",
        "weeks": delivery_columns(),
        "rows": [{"product": "R902745116", "orderQuantity": 5760, "weeklyQuantities": [1920, 1920, 0, 0, 1920, 0]}],
    })
    workbook = load_workbook(BytesIO(content), data_only=False)
    sheet = workbook["Teslimat Planı"]
    assert sheet["C4"].value == "Üretim"
    assert sheet["B7"].value == "R902745116"
    assert sheet["C7"].value == 5760
    assert sheet["D7"].value == "=SUM(I7:K7)"
    assert sheet["E7"].value == "=SUM(L7:N7)"
    assert sheet["F7"].value == "=D7+E7-C7"
    assert sheet["H7"].value == "R902745116"
    assert [sheet.cell(7, column).value for column in range(9, 15)] == [1920, 1920, 0, 0, 1920, 0]
    assert sheet["B50"].value == "Genel Toplam"
    assert sheet["I50"].value == "=SUM(I7:I49)"
    assert sheet["N50"].value == "=SUM(N7:N49)"
    assert sheet["I50"].fill.fgColor.rgb == "00FFF200"
    assert sheet["I50"].font.color.rgb == "00000000"
    assert sheet["I50"].font.bold is True
    assert sheet["P4"].value is None
    assert not list(workbook.defined_names)
    with ZipFile(BytesIO(content)) as archive:
        workbook_xml = archive.read("xl/workbook.xml").decode("utf-8")
    assert "#REF!" not in workbook_xml


def test_delivery_plan_endpoint_returns_an_excel_download():
    payload = {
        "startDate": "2026-07-20", "endDate": "2026-08-30", "category": "Üretim",
        "weeks": delivery_columns(),
        "rows": [{"product": "R902745116", "orderQuantity": 5760, "weeklyQuantities": [1920, 1920, 0, 0, 1920, 0]}],
    }
    with TestClient(app) as client:
        assert client.post("/api/delivery-plan/export", json=payload).status_code == 401
        login = client.post("/api/auth/login", json={"password": "test-password"})
        headers = {"Authorization": f"Bearer {login.json()['token']}"}
        response = client.post("/api/delivery-plan/export", headers=headers, json=payload)
        assert response.status_code == 200
        assert response.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        assert response.content.startswith(b"PK")


def test_delivery_plan_expands_when_more_than_template_rows_are_active():
    template = Path(__file__).resolve().parents[1] / "teslimat_plani.xlsx"
    rows = [{"product": f"P{index:03d}", "orderQuantity": index, "weeklyQuantities": [index, 0, 0, 0, 0, 0]} for index in range(1, 51)]
    content = build_delivery_plan(template, {
        "startDate": "2026-07-20", "endDate": "2026-08-30", "category": "Üretim",
        "weeks": delivery_columns(), "rows": rows,
    })
    sheet = load_workbook(BytesIO(content), data_only=False)["Teslimat Planı"]
    assert sheet["B56"].value == "P050"
    assert sheet["B57"].value == "Genel Toplam"
    assert sheet["C57"].value == "=SUM(C7:C56)"
    assert sheet["F57"].value == "=D57+E57-C57"
    assert sheet["I57"].value == "=SUM(I7:I56)"


def test_delivery_plan_expands_week_columns_beyond_the_original_six():
    template = Path(__file__).resolve().parents[1] / "teslimat_plani.xlsx"
    quantities = [100] * 9
    content = build_delivery_plan(template, {
        "startDate": "2026-06-22", "endDate": "2026-08-23", "category": "Üretim",
        "weeks": delivery_columns(9),
        "rows": [{"product": "P001", "orderQuantity": 900, "weeklyQuantities": quantities}],
    })
    sheet = load_workbook(BytesIO(content), data_only=False)["Teslimat Planı"]
    assert sheet["Q6"].value == "KW37\nPlan"
    assert sheet["Q7"].value == 100
    assert sheet["D7"].value == "=SUM(I7:K7)"
    assert sheet["E7"].value == "=SUM(L7:Q7)"
    assert sheet["Q50"].value == "=SUM(Q7:Q49)"
    assert sheet.auto_filter.ref == "B6:Q49"
