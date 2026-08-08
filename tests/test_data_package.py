import os
from io import BytesIO
from pathlib import Path

os.environ["DATABASE_PATH"] = "/tmp/selsa-planlama-feedback-test.sqlite"
os.environ["UPLOAD_DIR"] = "/tmp/selsa-planlama-feedback-uploads"
os.environ["ADMIN_PASSWORD"] = "test-password"
os.environ["APP_SESSION_SECRET"] = "test-session-secret-that-is-long-enough"

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.data_package import DataPackageError, build_data_package, parse_data_package
from app.main import app


def sample_parameters():
    return {
        "products": [{"product": "R1", "setupFamily": "piston", "batchSize": 1920, "dailyRate": 375, "shiftRate": 125, "parameterDailyRate": 375, "diameter": "17,2", "source": "Test"}],
        "preferences": [{"key": "R1", "machines": ["C-01", "C-02"]}],
        "restrictions": [],
        "machineRates": [{"id": "C-01", "rates": {"R1": 375}}],
    }


def test_data_package_round_trip_and_visible_sheets():
    data = sample_parameters()
    content = build_data_package("parameters", data)
    workbook = load_workbook(BytesIO(content), read_only=True)
    assert {"Bilgi", "Ürün Parametreleri", "Tezgah Öncelikleri", "__SelsaVeri"}.issubset(workbook.sheetnames)
    assert workbook["Ürün Parametreleri"]["B2"].value == "Piston"
    assert parse_data_package(content, "parameters") == data


def test_data_package_rejects_wrong_scope():
    content = build_data_package("parameters", sample_parameters())
    try:
        parse_data_package(content, "settings")
        assert False, "wrong scope should fail"
    except DataPackageError as error:
        assert "Ayarlar" in str(error)


def test_data_package_api_is_authenticated_and_round_trips():
    Path(os.environ["DATABASE_PATH"]).unlink(missing_ok=True)
    with TestClient(app) as client:
        payload = {"scope": "parameters", "data": sample_parameters()}
        assert client.post("/api/data-packages/export", json=payload).status_code == 401
        login = client.post("/api/auth/login", json={"password": "test-password"})
        headers = {"Authorization": f"Bearer {login.json()['token']}"}
        exported = client.post("/api/data-packages/export", headers=headers, json=payload)
        assert exported.status_code == 200
        imported = client.post(
            "/api/data-packages/parameters/import",
            headers=headers,
            files={"file": ("Selsa_Parametreler.xlsx", exported.content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert imported.status_code == 200
        assert imported.json()["data"] == sample_parameters()
