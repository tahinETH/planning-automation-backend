from io import BytesIO

from openpyxl import load_workbook

from app.overview_export import build_overview_workbook


def test_general_overview_export_contains_colored_summary_and_machine_tables():
    payload = {
            "generatedAt": "2026-07-30T10:00:00Z",
            "createdAt": "2026-07-30T09:30:00Z",
            "planState": "planned",
            "dirty": False,
            "summary": {
                "demand": 1000,
                "planned": 850,
                "plannedBatchCount": 3,
                "unplannedBatchCount": 1,
                "activeMachineCount": 2,
                "machineCount": 2,
                "machinesWithWorkCount": 2,
                "plannedMachineCount": 2,
            },
            "machines": [
                {
                    "id": "C-01",
                    "name": "CITIZEN 1",
                    "active": True,
                    "plannedCount": 1,
                    "plannedQuantity": 500,
                    "rows": [
                        {"kind": "current", "position": 0, "product": "R902740", "diameter": "23", "quantity": 350, "endDate": "2026-08-01", "workOrder": "320-1"},
                        {"kind": "planned", "position": 1, "product": "R902121", "diameter": "23", "quantity": 500, "endDate": "2026-08-04", "workOrder": ""},
                    ],
                },
                {
                    "id": "C-02",
                    "name": "CITIZEN 2",
                    "active": True,
                    "plannedCount": 2,
                    "plannedQuantity": 350,
                    "rows": [{"kind": "planned", "position": 1, "product": "R902690", "diameter": "20", "quantity": 350, "endDate": "2026-08-06", "workOrder": ""}],
                },
            ],
            "findings": [{"severity": "critical", "title": "Eksik üretim", "detail": "150 adet eksik.", "target": "Siparişler"}],
    }

    workbook = load_workbook(BytesIO(build_overview_workbook(payload)))
    assert workbook.sheetnames == ["Genel Bakış", "Plan Kontrolü"]
    sheet = workbook["Genel Bakış"]
    assert sheet["A1"].value == "SELSA  ·  ÜRETİM GENEL BAKIŞ"
    assert sheet["A6"].value == 1000
    assert sheet["G6"].value == 150
    assert sheet["A10"].value.startswith("C-01")
    assert sheet["A13"].value == "Üretimde"
    assert sheet["A13"].fill.fgColor.rgb.endswith("E4F5E4")
    assert workbook["Plan Kontrolü"]["A5"].value == "Kritik"
