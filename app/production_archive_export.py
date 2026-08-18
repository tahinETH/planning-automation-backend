from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


NAVY = "315C9B"
NAVY_LIGHT = "E8EEF7"
GREEN = "1E7B57"
GREEN_LIGHT = "EAF5EE"
INK = "18201D"
MUTED = "6C756F"
LINE = "DDE2DC"
PAPER = "F7F8F5"
WHITE = "FFFFFF"


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def build_production_archive_workbook(payload: dict[str, Any]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Üretim Arşivi"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A9"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    rows = payload.get("rows", [])
    status = payload.get("status", "semi-finished")
    status_label = "Yarı mamuller" if status == "semi-finished" else "Iskartalar" if status == "scrapped" else "Teslim edilenler"
    generated_at = payload.get("generatedAt") or datetime.now(timezone.utc).isoformat()
    total_quantity = sum(max(0, int(row.get("completedQuantity", 0))) for row in rows)
    machine_count = len({row.get("machineId") for row in rows if row.get("machineId")})

    sheet.merge_cells("A1:M2")
    title = sheet["A1"]
    title.value = "SELSA  ·  ÜRETİM ARŞİVİ"
    title.fill = _fill(NAVY)
    title.font = Font(name="Aptos Display", size=18, color=WHITE, bold=True)
    title.alignment = Alignment(vertical="center")
    for row in sheet["A1:M2"]:
        for cell in row:
            cell.fill = _fill(NAVY)

    sheet.merge_cells("A3:F3")
    sheet["A3"] = f"Durum: {status_label}"
    sheet["A3"].font = Font(name="Aptos", size=9, color=GREEN, bold=True)
    sheet.merge_cells("G3:M3")
    sheet["G3"] = f"Oluşturulma: {generated_at}"
    sheet["G3"].font = Font(name="Aptos", size=8, color=MUTED)
    sheet["G3"].alignment = Alignment(horizontal="right")

    metrics = [
        ("KAYIT", len(rows), "görünen satır"),
        ("TOPLAM ISKARTA" if status == "scrapped" else "TOPLAM ÜRETİM", total_quantity, "adet"),
        ("TEZGAH", machine_count, "farklı tezgah"),
        ("KAPSAM", "Filtreli" if payload.get("filtered") else "Tümü", f"{payload.get('totalAvailable', len(rows))} kayıt içinde"),
    ]
    for index, (label, value, detail) in enumerate(metrics):
        start_col = 1 + index * 2
        end_col = start_col + (3 if index == 3 else 1)
        start = get_column_letter(start_col)
        end = get_column_letter(end_col)
        sheet.merge_cells(f"{start}5:{end}5")
        sheet.merge_cells(f"{start}6:{end}6")
        sheet.merge_cells(f"{start}7:{end}7")
        sheet[f"{start}5"] = label
        sheet[f"{start}6"] = value
        sheet[f"{start}7"] = detail
        sheet[f"{start}5"].fill = _fill(NAVY_LIGHT)
        sheet[f"{start}5"].font = Font(name="Aptos", size=8, color=MUTED, bold=True)
        sheet[f"{start}6"].fill = _fill(WHITE)
        sheet[f"{start}6"].font = Font(name="Aptos Display", size=16, color=GREEN if index < 3 else NAVY, bold=True)
        sheet[f"{start}7"].fill = _fill(WHITE)
        sheet[f"{start}7"].font = Font(name="Aptos", size=7, color=MUTED)
        if isinstance(value, (int, float)):
            sheet[f"{start}6"].number_format = "#,##0"

    headings = [str(payload.get("eventDateLabel") or "Tamamlanma"), "Tezgah", "Tezgah adı", "İş emri", "Ürün", "Ürün ailesi", "Tamamlanan proses", "Mevcut aşama", "Sıradaki operasyon", "Iskarta" if status == "scrapped" else "Üretilen", "Planlanan başlangıç", "Planlanan bitiş", "Teslim tarihi"]
    widths = [16, 12, 24, 18, 20, 15, 17, 23, 18, 14, 20, 20, 16]
    thin = Side(style="thin", color=LINE)
    for column, (heading, width) in enumerate(zip(headings, widths, strict=True), 1):
        cell = sheet.cell(9, column, heading)
        cell.fill = _fill(NAVY)
        cell.font = Font(name="Aptos", size=8, color=WHITE, bold=True)
        cell.alignment = Alignment(vertical="center", horizontal="right" if column == 10 else "left")
        cell.border = Border(bottom=thin)
        sheet.column_dimensions[get_column_letter(column)].width = width

    process_labels = {"turning": "Torna", "drilling": "Delme", "deburring": "Çapak alma", "washing": "Yıkama", "gkm": "GKM"}
    family_labels = {"piston": "Piston", "center-pin": "Center pim"}
    for row_index, item in enumerate(rows, 10):
        values = [
            item.get("completedAt", ""),
            item.get("machineId", ""),
            item.get("machineName", ""),
            item.get("workOrder", ""),
            item.get("product", ""),
            family_labels.get(item.get("setupFamily"), "Aile tanımsız"),
            process_labels.get(item.get("process"), item.get("process", "")),
            item.get("stage", ""),
            item.get("currentOperation", ""),
            max(0, int(item.get("completedQuantity", 0))),
            item.get("plannedStart", ""),
            item.get("plannedEnd", ""),
            item.get("deliveryDate", ""),
        ]
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row_index, column, value)
            cell.fill = _fill(WHITE if row_index % 2 == 0 else PAPER)
            cell.font = Font(name="Aptos", size=8, color=INK, bold=column in {2, 5})
            cell.alignment = Alignment(vertical="center", horizontal="right" if column == 10 else "left")
            cell.border = Border(bottom=thin)
            if column == 10:
                cell.number_format = "#,##0"

    last_row = max(9, sheet.max_row)
    sheet.auto_filter.ref = f"A9:M{last_row}"
    sheet.row_dimensions[1].height = 25
    sheet.row_dimensions[9].height = 25

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
