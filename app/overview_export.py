from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


NAVY = "315C9B"
NAVY_LIGHT = "E8EEF7"
PEACH = "F1D0BC"
GREEN = "1E7B57"
GREEN_LIGHT = "E4F5E4"
ORANGE_LIGHT = "FFF0E9"
RED = "C84C43"
RED_LIGHT = "FFF0EE"
AMBER = "A86C24"
AMBER_LIGHT = "FFF4DF"
INK = "18201D"
MUTED = "6C756F"
LINE = "DDE2DC"
PAPER = "F6F7F3"
WHITE = "FFFFFF"

THIN_LINE = Side(style="thin", color=LINE)


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _merge_value(sheet, cell_range: str, value: Any, **style: Any) -> None:
    sheet.merge_cells(cell_range)
    cell = sheet[cell_range.split(":")[0]]
    cell.value = value
    if "fill" in style:
        cell.fill = _fill(style["fill"])
    if "font" in style:
        cell.font = style["font"]
    if "alignment" in style:
        cell.alignment = style["alignment"]
    if "border" in style:
        cell.border = style["border"]


def _kpi(sheet, start_col: int, end_col: int, label: str, value: Any, detail: str, *, tone: str = NAVY) -> None:
    start = get_column_letter(start_col)
    end = get_column_letter(end_col)
    _merge_value(
        sheet,
        f"{start}5:{end}5",
        label,
        fill=NAVY_LIGHT if tone == NAVY else (GREEN_LIGHT if tone == GREEN else RED_LIGHT),
        font=Font(name="Aptos", size=9, color=MUTED, bold=True),
        alignment=Alignment(vertical="center"),
    )
    _merge_value(
        sheet,
        f"{start}6:{end}7",
        value,
        fill=WHITE,
        font=Font(name="Aptos Display", size=20, color=tone, bold=True),
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    if isinstance(value, (int, float)):
        sheet[f"{start}6"].number_format = "#,##0"
    _merge_value(
        sheet,
        f"{start}8:{end}8",
        detail,
        fill=WHITE,
        font=Font(name="Aptos", size=8, color=MUTED),
        alignment=Alignment(vertical="center"),
    )
    for row in range(5, 9):
        for col in range(start_col, end_col + 1):
            sheet.cell(row, col).border = Border(bottom=THIN_LINE, left=THIN_LINE if col == start_col else None, right=THIN_LINE if col == end_col else None)


def _machine_block(sheet, machine: dict[str, Any], start_row: int, start_col: int, height: int) -> None:
    end_col = start_col + 6
    first = get_column_letter(start_col)
    last = get_column_letter(end_col)
    active = bool(machine.get("active"))
    header_color = NAVY if active else MUTED
    status = "Aktif" if active else "Pasif"

    _merge_value(
        sheet,
        f"{first}{start_row}:{last}{start_row + 1}",
        f"{machine.get('id', '')}  ·  {machine.get('name', '')}     {status}",
        fill=header_color,
        font=Font(name="Aptos Display", size=12, color=WHITE, bold=True),
        alignment=Alignment(vertical="center"),
    )
    headings = ["Sıra", "Ürün", "Çap", "Adet", "Bitiş", "İş emri", "Durum"]
    for offset, heading in enumerate(headings):
        cell = sheet.cell(start_row + 2, start_col + offset, heading)
        cell.fill = _fill(PEACH)
        cell.font = Font(name="Aptos", size=8, color="56301E", bold=True)
        cell.alignment = Alignment(horizontal="center" if offset in {0, 2, 3, 4, 6} else "left", vertical="center")
        cell.border = Border(bottom=THIN_LINE)

    rows = machine.get("rows", [])
    data_height = max(1, height)
    for index in range(data_height):
        row_number = start_row + 3 + index
        item = rows[index] if index < len(rows) else None
        kind = item.get("kind") if item else ""
        background = GREEN_LIGHT if kind == "current" else (WHITE if index % 2 == 0 else PAPER)
        values = (
            [
                "Üretimde" if kind == "current" else item.get("position", ""),
                item.get("product", ""),
                item.get("diameter", ""),
                item.get("quantity", 0),
                item.get("endDate", ""),
                item.get("workOrder", ""),
                "Mevcut iş" if kind == "current" else "Planlı",
            ]
            if item
            else ["—", "Kuyruk boş", "", "", "", "", ""]
        )
        for offset, value in enumerate(values):
            cell = sheet.cell(row_number, start_col + offset, value)
            cell.fill = _fill(background)
            cell.font = Font(name="Aptos", size=8, color=GREEN if kind == "current" and offset == 6 else INK, bold=offset in {1, 6})
            cell.alignment = Alignment(horizontal="center" if offset in {0, 2, 3, 4, 6} else "left", vertical="center")
            cell.border = Border(bottom=THIN_LINE)
            if offset == 3 and isinstance(value, (int, float)):
                cell.number_format = "#,##0"

    footer_row = start_row + 3 + data_height
    _merge_value(
        sheet,
        f"{first}{footer_row}:{last}{footer_row}",
        f"{machine.get('plannedCount', 0)} planlı şarj  ·  {machine.get('plannedQuantity', 0):,.0f} adet",
        fill="FBFCFA",
        font=Font(name="Aptos", size=8, color=MUTED, bold=True),
        alignment=Alignment(horizontal="right", vertical="center"),
    )
    for row in range(start_row, footer_row + 1):
        sheet.cell(row, start_col).border = Border(left=THIN_LINE, bottom=sheet.cell(row, start_col).border.bottom)
        sheet.cell(row, end_col).border = Border(right=THIN_LINE, bottom=sheet.cell(row, end_col).border.bottom)


def _audit_sheet(workbook: Workbook, findings: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("Plan Kontrolü")
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A4"
    _merge_value(
        sheet,
        "A1:D2",
        "PLAN KONTROLÜ",
        fill=NAVY,
        font=Font(name="Aptos Display", size=17, color=WHITE, bold=True),
        alignment=Alignment(vertical="center"),
    )
    for column, width in {"A": 15, "B": 34, "C": 88, "D": 24}.items():
        sheet.column_dimensions[column].width = width
    headers = ["Seviye", "Kontrol", "Açıklama", "İlgili alan"]
    for index, value in enumerate(headers, 1):
        cell = sheet.cell(4, index, value)
        cell.fill = _fill(PEACH)
        cell.font = Font(name="Aptos", size=9, color="56301E", bold=True)
        cell.border = Border(bottom=THIN_LINE)
    if not findings:
        sheet.append(["Bilgi", "Temel kontroller temiz", "Plan kontrolünde açık bulgu bulunmadı.", ""])
    else:
        for finding in findings:
            sheet.append([
                {"critical": "Kritik", "warning": "Uyarı", "info": "Bilgi"}.get(finding.get("severity"), "Bilgi"),
                finding.get("title", ""),
                finding.get("detail", ""),
                finding.get("target", ""),
            ])
    for row in range(5, sheet.max_row + 1):
        level = sheet.cell(row, 1).value
        tone, background = (RED, RED_LIGHT) if level == "Kritik" else ((AMBER, AMBER_LIGHT) if level == "Uyarı" else (NAVY, NAVY_LIGHT))
        for col in range(1, 5):
            cell = sheet.cell(row, col)
            cell.fill = _fill(background if col == 1 else WHITE)
            cell.font = Font(name="Aptos", size=9, color=tone if col == 1 else INK, bold=col in {1, 2})
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=THIN_LINE)
        sheet.row_dimensions[row].height = 36
    sheet.auto_filter.ref = f"A4:D{sheet.max_row}"


def build_overview_workbook(payload: dict[str, Any]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Genel Bakış"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A10"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    for column in range(1, 16):
        sheet.column_dimensions[get_column_letter(column)].width = 12
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["F"].width = 17
    sheet.column_dimensions["J"].width = 18
    sheet.column_dimensions["N"].width = 17
    sheet.column_dimensions["H"].width = 3

    _merge_value(
        sheet,
        "A1:O2",
        "SELSA  ·  ÜRETİM GENEL BAKIŞ",
        fill=NAVY,
        font=Font(name="Aptos Display", size=18, color=WHITE, bold=True),
        alignment=Alignment(vertical="center"),
    )
    generated = payload.get("generatedAt") or datetime.now(timezone.utc).isoformat()
    state = "Girdiler değişti · plan güncel değil" if payload.get("dirty") else ("Planlandı" if payload.get("planState") == "planned" else "Plan henüz çalıştırılmadı")
    _merge_value(sheet, "A3:H3", state, fill=WHITE, font=Font(name="Aptos", size=9, color=RED if payload.get("dirty") else GREEN, bold=True), alignment=Alignment(vertical="center"))
    _merge_value(sheet, "I3:O3", f"Oluşturulma: {generated}", fill=WHITE, font=Font(name="Aptos", size=8, color=MUTED), alignment=Alignment(horizontal="right", vertical="center"))

    summary = payload.get("summary", {})
    demand = max(0, int(summary.get("demand", 0)))
    planned = max(0, int(summary.get("planned", 0)))
    shortage = max(0, demand - planned)
    coverage = planned / demand if demand else 0
    _kpi(sheet, 1, 3, "SİPARİŞ TALEBİ", demand, "adet", tone=NAVY)
    _kpi(sheet, 4, 6, "KARŞILANAN", planned if payload.get("planState") == "planned" else "—", f"%{coverage * 100:.1f} karşılama" if payload.get("planState") == "planned" else "hesaplanmadı", tone=GREEN)
    _kpi(sheet, 7, 9, "EKSİK ÜRETİM", shortage if payload.get("planState") == "planned" else "—", f"{summary.get('unplannedBatchCount', 0)} planlanamayan şarj" if payload.get("planState") == "planned" else "hesaplanmadı", tone=RED)
    _kpi(sheet, 10, 12, "PLANLI ŞARJ", summary.get("plannedBatchCount", 0) if payload.get("planState") == "planned" else "—", f"{summary.get('plannedMachineCount', 0)} tezgaha dağıtıldı", tone=NAVY)
    _kpi(sheet, 13, 15, "AKTİF TEZGAH", f"{summary.get('activeMachineCount', 0)} / {summary.get('machineCount', 0)}", f"{summary.get('machinesWithWorkCount', 0)} tezgah iş gösteriyor", tone=NAVY)

    machines = payload.get("machines", [])
    current_row = 10
    for index in range(0, len(machines), 2):
        pair = machines[index:index + 2]
        height = max(1, *(len(machine.get("rows", [])) for machine in pair))
        _machine_block(sheet, pair[0], current_row, 1, height)
        if len(pair) > 1:
            _machine_block(sheet, pair[1], current_row, 9, height)
        current_row += height + 6

    sheet.print_area = f"A1:O{max(10, current_row - 2)}"
    sheet.auto_filter.ref = None
    _audit_sheet(workbook, payload.get("findings", []))

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
