from __future__ import annotations

import json
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Literal

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


DataPackageScope = Literal["parameters", "settings", "scenarios"]
MAX_DATA_PACKAGE_BYTES = 20 * 1024 * 1024
FORMAT_NAME = "SELSA_PLANLAMA_DATA_PACKAGE"
FORMAT_VERSION = 1
CHUNK_SIZE = 30_000

SCOPE_LABELS: dict[DataPackageScope, str] = {
    "parameters": "Parametreler",
    "settings": "Ayarlar",
    "scenarios": "Senaryolar",
}


class DataPackageError(ValueError):
    pass


def _header(sheet, values: list[str]) -> None:
    sheet.append(values)
    fill = PatternFill("solid", fgColor="24443D")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _append_rows(sheet, headers: list[str], rows: list[list[Any]]) -> None:
    _header(sheet, headers)
    for row in rows:
        sheet.append(row)
    for column in sheet.columns:
        width = min(45, max(10, max(len(str(cell.value or "")) for cell in column) + 2))
        sheet.column_dimensions[column[0].column_letter].width = width


def _parameter_sheets(workbook: Workbook, data: dict[str, Any]) -> None:
    products = data.get("products", [])
    parameter_sheet = workbook.create_sheet("Ürün Parametreleri")
    _append_rows(parameter_sheet, ["Ürün", "Şarj Büyüklüğü", "Günlük Üretim", "Vardiya Üretimi", "Parametre Üretimi", "Çap", "Kaynak"], [
        [
            item.get("product", ""),
            item.get("batchSize", 0),
            item.get("dailyRate", 0),
            item.get("shiftRate", 0),
            item.get("parameterDailyRate", 0),
            item.get("diameter", ""),
            item.get("source", ""),
        ]
        for item in products
    ])
    priority_sheet = workbook.create_sheet("Tezgah Öncelikleri")
    rows: list[list[Any]] = []
    machine_rates = {item.get("id"): item.get("rates", {}) for item in data.get("machineRates", [])}
    for product in products:
        machines = next((rule.get("machines", []) for rule in data.get("preferences", []) if rule.get("key") == product.get("product")), product.get("eligibleMachines", []))
        for index, machine_id in enumerate(machines):
            rows.append([product.get("product", ""), index + 1, machine_id, machine_rates.get(machine_id, {}).get(str(product.get("product", "")).upper(), "")])
    _append_rows(priority_sheet, ["Ürün", "Öncelik", "Tezgah", "Günlük Üretim"], rows)


def _settings_sheets(workbook: Workbook, data: dict[str, Any]) -> None:
    setup = data.get("setupSettings", {}) or {}
    setup_sheet = workbook.create_sheet("Genel Ayarlar")
    _append_rows(setup_sheet, ["Ayar", "Değer"], [
        ["Vardiya süresi (saat)", setup.get("shiftHours", "")],
        ["Aynı çap setup (saat)", setup.get("sameDiameterHours", "")],
        ["Farklı çap setup (saat)", setup.get("differentDiameterHours", "")],
    ])
    holiday_sheet = workbook.create_sheet("Tatil Günleri")
    _append_rows(holiday_sheet, ["Tatil", "Excel Tarih Seri No", "Çalışılan Vardiya"], [
        [item.get("name", ""), item.get("serial", ""), item.get("workingShifts", 0)]
        for item in data.get("holidays", [])
    ])
    event_sheet = workbook.create_sheet("Tezgah Takvimi")
    _append_rows(event_sheet, ["Kayıt ID", "Tezgah", "Tür", "Açıklama", "Başlangıç", "Bitiş", "Vardiya"], [
        [item.get("id", ""), item.get("machineId", ""), item.get("kind", ""), item.get("name", ""), item.get("start", ""), item.get("end", ""), item.get("shiftCount", 0)]
        for item in data.get("calendarEvents", [])
    ])
    machine_sheet = workbook.create_sheet("Tezgah Ayarları")
    _append_rows(machine_sheet, ["Tezgah", "Ad", "Aktif", "Müsait Başlangıç", "Plan Bitişi", "Vardiya", "Kapasite"], [
        [item.get("id", ""), item.get("name", ""), item.get("active", False), item.get("availableStart", ""), item.get("planEnd", ""), item.get("shiftFactor", 0), item.get("capacityFactor", 1)]
        for item in data.get("machines", [])
    ])


def _scenario_sheets(workbook: Workbook, data: Any) -> None:
    scenarios = data if isinstance(data, list) else data.get("scenarios", [])
    sheet = workbook.create_sheet("Senaryolar")
    _append_rows(sheet, ["Senaryo", "Açıklama", "Oluşturulma", "Talep", "Planlanan", "Şarj", "Plan Bitişi"], [
        [
            item.get("name", ""),
            item.get("notes", ""),
            item.get("createdAt", ""),
            item.get("result", {}).get("summary", {}).get("demand", 0),
            item.get("result", {}).get("summary", {}).get("planned", 0),
            item.get("result", {}).get("summary", {}).get("plannedBatchCount", 0),
            item.get("result", {}).get("summary", {}).get("completion", ""),
        ]
        for item in scenarios
    ])


def build_data_package(scope: DataPackageScope, data: Any) -> bytes:
    workbook = Workbook()
    info = workbook.active
    info.title = "Bilgi"
    info.append(["Selsa Planlama Excel Veri Paketi"])
    info["A1"].font = Font(size=16, bold=True, color="24443D")
    info.append(["Kapsam", SCOPE_LABELS[scope]])
    info.append(["Format sürümü", FORMAT_VERSION])
    info.append(["Oluşturulma", datetime.now(timezone.utc).isoformat()])
    info.append(["Not", "Bu dosya Selsa Planlama uygulamasına tekrar yüklenebilir. Gizli veri sayfasını silmeyin."])
    info.column_dimensions["A"].width = 24
    info.column_dimensions["B"].width = 72

    if scope == "parameters":
        _parameter_sheets(workbook, data)
    elif scope == "settings":
        _settings_sheets(workbook, data)
    else:
        _scenario_sheets(workbook, data)

    payload = json.dumps({"format": FORMAT_NAME, "version": FORMAT_VERSION, "scope": scope, "data": data}, ensure_ascii=False, separators=(",", ":"))
    storage = workbook.create_sheet("__SelsaVeri")
    storage.append(["Sıra", "Veri"])
    for index in range(0, len(payload), CHUNK_SIZE):
        storage.append([index // CHUNK_SIZE + 1, payload[index:index + CHUNK_SIZE]])
    storage.sheet_state = "veryHidden"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def parse_data_package(content: bytes, expected_scope: DataPackageScope) -> Any:
    if not content:
        raise DataPackageError("Excel dosyası boş.")
    if len(content) > MAX_DATA_PACKAGE_BYTES:
        raise DataPackageError("Excel veri paketi 20 MB sınırını aşıyor.")
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise DataPackageError("Dosya geçerli bir .xlsx çalışma kitabı değil.") from None
    if "__SelsaVeri" not in workbook.sheetnames:
        raise DataPackageError("Bu dosya Selsa Planlama tarafından oluşturulmuş bir veri paketi değil.")
    rows = list(workbook["__SelsaVeri"].iter_rows(min_row=2, values_only=True))
    chunks = [(int(row[0]), str(row[1] or "")) for row in rows if row[0] is not None]
    if not chunks:
        raise DataPackageError("Excel veri paketindeki uygulama verisi bulunamadı.")
    try:
        payload = json.loads("".join(value for _, value in sorted(chunks)))
    except (TypeError, ValueError):
        raise DataPackageError("Excel veri paketindeki uygulama verisi okunamadı.") from None
    if payload.get("format") != FORMAT_NAME or payload.get("version") != FORMAT_VERSION:
        raise DataPackageError("Excel veri paketinin sürümü desteklenmiyor.")
    if payload.get("scope") != expected_scope:
        raise DataPackageError(f"Bu dosya {SCOPE_LABELS[expected_scope]} alanına ait değil.")
    return payload.get("data")
