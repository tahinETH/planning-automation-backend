from __future__ import annotations

from copy import copy
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

MAX_PRODUCTS = 200
MAX_WEEKS = 55
TEMPLATE_PRODUCT_ROWS = 43
TEMPLATE_WEEK_COLUMNS = 6


class DeliveryPlanError(ValueError):
    pass


def _parse_date(value: str, label: str) -> date:
    try:
        return date.fromisoformat(value)
    except (TypeError, ValueError) as error:
        raise DeliveryPlanError(f"{label} geçerli bir tarih olmalıdır.") from error


def _sum_formula(row_index: int, column_indexes: list[int]) -> str:
    if not column_indexes:
        return "=0"
    cells = [f"{get_column_letter(column)}{row_index}" for column in column_indexes]
    if column_indexes == list(range(column_indexes[0], column_indexes[-1] + 1)):
        return f"=SUM({cells[0]}:{cells[-1]})"
    return f"=SUM({','.join(cells)})"


def build_delivery_plan(template_path: Path, payload: dict) -> bytes:
    if not template_path.exists():
        raise DeliveryPlanError("Teslimat planı şablonu bulunamadı.")

    weeks = payload.get("weeks") or []
    rows = payload.get("rows") or []
    if not 1 <= len(weeks) <= MAX_WEEKS:
        raise DeliveryPlanError(f"Teslimat planı 1–{MAX_WEEKS} hafta içermelidir.")
    if len(rows) > MAX_PRODUCTS:
        raise DeliveryPlanError(f"Şablon en fazla {MAX_PRODUCTS} aktif sipariş destekliyor.")

    start_date = _parse_date(payload.get("startDate", ""), "Başlangıç tarihi")
    end_date = _parse_date(payload.get("endDate", ""), "Bitiş tarihi")
    if end_date < start_date:
        raise DeliveryPlanError("Bitiş tarihi başlangıç tarihinden önce olamaz.")

    workbook = load_workbook(template_path, keep_links=False)
    sheet = workbook["Teslimat Planı"]
    workbook.defined_names.clear()
    sheet.defined_names.clear()
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"

    if "P4:R4" in {str(item) for item in sheet.merged_cells.ranges}:
        sheet.unmerge_cells("P4:R4")
    sheet.delete_cols(14, 13)
    sheet.insert_cols(6, 1)
    for merged_range, shifted_range in (("H1:J1", "I1:K1"), ("H2:J2", "I2:K2")):
        if merged_range in {str(item) for item in sheet.merged_cells.ranges}:
            sheet.unmerge_cells(merged_range)
        sheet.merge_cells(shifted_range)
    for row_index in range(1, sheet.max_row + 1):
        source = sheet.cell(row_index, 5)
        target = sheet.cell(row_index, 6)
        target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)
    sheet.column_dimensions["F"].width = sheet.column_dimensions["E"].width

    extra_rows = max(0, len(rows) - TEMPLATE_PRODUCT_ROWS)
    if extra_rows:
        sheet.insert_rows(50, amount=extra_rows)
        for row_index in range(50, 50 + extra_rows):
            for column in range(1, 15):
                source = sheet.cell(49, column)
                target = sheet.cell(row_index, column)
                target._style = copy(source._style)
                target.number_format = source.number_format
                target.alignment = copy(source.alignment)
                target.protection = copy(source.protection)
    total_row = 50 + extra_rows
    summary_header_row = 53 + extra_rows
    summary_total_row = 54 + extra_rows

    extra_week_columns = max(0, len(weeks) - TEMPLATE_WEEK_COLUMNS)
    if extra_week_columns:
        sheet.insert_cols(15, amount=extra_week_columns)
        for column in range(15, 15 + extra_week_columns):
            for row_index in range(1, summary_total_row + 1):
                source = sheet.cell(row_index, 14)
                target = sheet.cell(row_index, column)
                target._style = copy(source._style)
                target.number_format = source.number_format
                target.alignment = copy(source.alignment)
                target.protection = copy(source.protection)
            sheet.column_dimensions[get_column_letter(column)].width = sheet.column_dimensions["N"].width
    weekly_end_column = 8 + len(weeks)
    weekly_end_letter = get_column_letter(weekly_end_column)
    actual_columns = [9 + offset for offset, week in enumerate(weeks) if week.get("kind") == "actual"]
    planned_columns = [9 + offset for offset, week in enumerate(weeks) if week.get("kind") == "plan"]

    sheet["C1"] = datetime.combine(datetime.now(ZoneInfo("Europe/Istanbul")).date(), datetime.min.time())
    sheet["H1"] = "Başlangıç T.:"
    sheet["H2"] = "Bitiş Tarihi"
    sheet["I1"] = datetime.combine(start_date, datetime.min.time())
    sheet["I2"] = datetime.combine(end_date, datetime.min.time())
    sheet["I1"].number_format = "m/d/yy"
    sheet["I2"].number_format = "m/d/yy"
    sheet["C4"] = str(payload.get("category") or "Üretim")
    sheet["B6"] = "Tip No"
    sheet["C6"] = "Sipariş Adeti"
    sheet["D6"] = "Teslimat Adeti"
    sheet["E6"] = "Plan Adet"
    sheet["F6"] = "Fark"
    sheet["H6"] = "Tip No"
    sheet.column_dimensions["H"].width = 18

    for offset in range(TEMPLATE_WEEK_COLUMNS):
        column = 9 + offset
        sheet.cell(6, column).value = weeks[offset]["label"] if offset < len(weeks) else None
    for offset in range(TEMPLATE_WEEK_COLUMNS, len(weeks)):
        sheet.cell(6, 9 + offset).value = weeks[offset]["label"]
    sheet.row_dimensions[6].height = 34
    for offset, week in enumerate(weeks):
        cell = sheet.cell(6, 9 + offset)
        sheet.column_dimensions[get_column_letter(9 + offset)].width = 14 if week.get("kind") != "carryover" else 15
        alignment = copy(cell.alignment)
        alignment.wrap_text = True
        cell.alignment = alignment
        font = copy(cell.font)
        font.sz = 9
        if week.get("kind") == "actual":
            font.color = "009A44"
        elif week.get("kind") == "carryover":
            font.color = "C15B26"
        cell.font = font

    for row_index in range(7, total_row):
        for column in (*range(2, 7), *range(8, weekly_end_column + 1)):
            sheet.cell(row_index, column).value = None

    for row_index, row in enumerate(rows, start=7):
        product = str(row.get("product") or "").strip().upper()
        if not product:
            raise DeliveryPlanError("Boş malzeme kodu dışa aktarılamaz.")
        quantities = row.get("weeklyQuantities") or []
        if len(quantities) != len(weeks):
            raise DeliveryPlanError(f"{product} için haftalık adet sayısı başlıklarla uyuşmuyor.")
        order_quantity = max(0, int(row.get("orderQuantity") or 0))
        sheet.cell(row_index, 2).value = product
        sheet.cell(row_index, 3).value = order_quantity
        sheet.cell(row_index, 4).value = _sum_formula(row_index, actual_columns)
        sheet.cell(row_index, 5).value = _sum_formula(row_index, planned_columns)
        sheet.cell(row_index, 6).value = f"=D{row_index}+E{row_index}-C{row_index}"
        sheet.cell(row_index, 8).value = product
        for offset in range(len(weeks)):
            value_cell = sheet.cell(row_index, 9 + offset)
            value_cell.value = max(0, int(quantities[offset]))
            font = copy(value_cell.font)
            if weeks[offset].get("kind") == "actual":
                font.color = "009A44"
                font.bold = True
            elif weeks[offset].get("kind") == "carryover":
                font.color = "C15B26"
            value_cell.font = font

    sheet.cell(total_row, 2).value = "Genel Toplam"
    sheet.cell(total_row, 3).value = f"=SUM(C7:C{total_row - 1})"
    sheet.cell(total_row, 4).value = f"=SUM(D7:D{total_row - 1})"
    sheet.cell(total_row, 5).value = f"=SUM(E7:E{total_row - 1})"
    sheet.cell(total_row, 6).value = f"=D{total_row}+E{total_row}-C{total_row}"
    general_total_fill = PatternFill("solid", fgColor="FFF200")
    for column in range(9, weekly_end_column + 1):
        cell = sheet.cell(total_row, column)
        column_letter = get_column_letter(column)
        cell.value = f"=SUM({column_letter}7:{column_letter}{total_row - 1})"
        cell.fill = general_total_fill
        font = copy(cell.font)
        font.color = "000000"
        font.bold = True
        cell.font = font
    sheet.cell(total_row, 2).fill = general_total_fill
    total_label_font = copy(sheet.cell(total_row, 2).font)
    total_label_font.color = "000000"
    total_label_font.bold = True
    sheet.cell(total_row, 2).font = total_label_font

    sheet.cell(summary_header_row, 3).value = "Sipariş"
    sheet.cell(summary_header_row, 4).value = "Teslimat"
    sheet.cell(summary_header_row, 5).value = "Plan"
    sheet.cell(summary_header_row, 6).value = "Fark"
    sheet.cell(summary_total_row, 2).value = "Toplam Üretim"
    sheet.cell(summary_total_row, 3).value = f"=C{total_row}"
    sheet.cell(summary_total_row, 4).value = f"=D{total_row}"
    sheet.cell(summary_total_row, 5).value = f"=E{total_row}"
    sheet.cell(summary_total_row, 6).value = f"=F{total_row}"
    for row_index in (summary_total_row + 1, summary_total_row + 2):
        for column in range(2, 7):
            sheet.cell(row_index, column).value = None

    sheet.conditional_formatting._cf_rules.clear()
    sheet.conditional_formatting.add(f"F7:F{total_row}", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="FFC7CE")))
    sheet.conditional_formatting.add(f"F7:F{total_row}", CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="C6EFCE")))
    sheet.freeze_panes = "B7"
    sheet.auto_filter.ref = f"B6:{weekly_end_letter}{total_row - 1}"
    sheet.print_area = f"B1:{weekly_end_letter}{summary_total_row}"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.page_setup.orientation = "landscape"
    sheet.sheet_view.zoomScale = 90

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
